import os
from datetime import datetime

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

if os.path.exists('agırlıktakibi.xlsx'):
    wb = load_workbook('agırlıktakibi.xlsx')
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    wb.save(filename='agırlıktakibi.xlsx')
    ws.append(["Tarih", "Agırlık(kg)"])
    wb.save()
agirliklar = []
tarih = datetime.today().date()
agirlik = float(input("Bugünkü ağırlığın ne (kg) ?"))
ws.append([tarih, agirlik])
wb.save("agırlıktakibi.xlsx")
for row in ws.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
    kilo = row[0]
    if kilo is not None:
        agirliklar.append(kilo)
ortalama = sum(agirliklar) / len(agirliklar)
minimum = min(agirliklar)
maksimum = max(agirliklar)

print(f"Ağırlık ortalaması: {ortalama:.2f} kg")
print(f"En düşük ağırlık: {minimum} kg")
print(f"En yüksek ağırlık: {maksimum} kg")
